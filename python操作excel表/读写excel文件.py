import openpyxl

#创建一个Excel workbook对象
book = openpyxl.workbook.Workbook()

#创建时，会自动产生一个sheet，通过active获取
sh = book.active

#修改当前sheet标题为工资表
sh.title = "工资表"

#保存文件
# book.save("信息.xlsx")

#增加一个名为“年龄表”的sheet，放在最后
sh1 = book.create_sheet('年龄表-最后')

# 增加一个 sheet，放在最前
sh2 = book.create_sheet('年龄表-最前',0)

# 增加一个 sheet，指定为第2个表单
sh3 = book.create_sheet('年龄表2',1)

#根据名称获取某个sheet对象
chos_sht = book['工资表']

#为选中的sheet的第一个单元格写入内容
chos_sht['A1']='你好'

#获取某个单元格内容
print(sh["A1"].value)

#根据行号列好，给第一个单元格写入内容，
#注意和xlrd不同，是从1开始
chos_sht.cell(2,2).value = "白月黑羽"

#根据行号列号，获取某个单元格内容
print(chos_sht.cell(1,1).value)
#保存文件
book.save("信息.xlsx")