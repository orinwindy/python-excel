import openpyxl
from openpyxl.styles import Font,colors#字体的样式设置
from openpyxl.styles import PatternFill#设置背景色
from openpyxl.drawing.image import Image#插入图片

wb = openpyxl.load_workbook('income.xlsx')

sheet = wb['2017']

#直接对单元格进行赋值即可
sheet["A1"]="修改一下"

# sheet 对象的 insert_rows 和 insert_cols 方法，分别用来插入行和列

#在第二行的位置插入一行
# sheet.insert_rows(2)
#删除第二行
# sheet.delete_rows(2)

#在第二列的位置插入一列
# sheet.insert_cols(2)
#删除第二列
# sheet.delete_cols(2)

#在第2行的位置插入3行
# sheet.insert_rows(2,3)
#在第2列的位置插入3行列
# sheet.insert_cols(2,3)

#文字、颜色、字体、大小的设置,均是通过font对象设定的
#指定单元格字体颜色，
sheet['A1'].font = Font(color=colors.BLUE,#使用预设的颜色常量
                        size=15,#设定文字大小
                        bold = True,#设定为粗体
                        italic = True#设定为斜体
                        )

#也可以用RGB数字表示颜色
sheet["B1"].font = Font(color="981818")

#指定整行的字体风格，整理指定的是第3行
font = Font(color='981818')
for y in range(1,100):
    sheet.cell(row=3,column=y).font =font

#指定整列 字体风格， 这里指定的是第2列
font = Font(bold=True)
for x in range(1, 100): # 第 1 到 100 行
    sheet.cell(row=x, column=2).font = font

#背景色设置
# 指定 某个单元格背景色
sheet["A1"].fill = PatternFill("solid","E39191")

# 指定 整行 背景色， 这里指定的是第2行
fill = PatternFill("solid","E39191")
for y in range (1,100):
    sheet.cell(row=2,column=y).fill = fill

#插入图片
# 在第1行，第4列 的位置插入图片
sheet.add_image(Image('1.jpg'), 'D1')

#保存文件，可另存
wb.save('income-1.xlsx')
